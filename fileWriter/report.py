
from nlp.nlp import posTagger
from openpyxl.styles import PatternFill, Border, Side, Alignment, alignment
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from mainProcess import MainProcess
from nlp.data.inputData import NNnlpInputEntry
from nlp.data.morph import Morph
import os

def draw_style(ws:Worksheet, row, col, option=0):
    if option == 2:
        fill_color = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    else:
        fill_color = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    align = Alignment(horizontal='center', vertical='center')
    ws.cell(row, col).fill = fill_color
    ws.cell(row, col).border = thin_border
    if option == 0:
        ws.cell(row, col).alignment = align

def write_report(inputData:NNnlpInputEntry, checkOptions:list) -> Workbook:

    load_wb = load_workbook(os.path.join(os.getcwd(),"docs/template.xlsx"))

    load_ws = load_wb["분석 결과 요약"]

    load_ws.cell(6,3,inputData.n_letters)
    load_ws.cell(7,3,inputData.n_lettersIncldSpc)
    load_ws.cell(8,3,inputData.n_morph)
    load_ws.cell(9,3,inputData.n_word)
    load_ws.cell(10,3,inputData.n_eojeol)
    load_ws.cell(11,3,inputData.n_jeol)
    load_ws.cell(12,3,inputData.n_sentences)
    load_ws.cell(13,3,inputData.n_paragraph)

    load_ws.cell(18,3,inputData.n_real_morph)
    load_ws.cell(19,3,inputData.n_grammar_morph)
    load_ws.cell(20,3,inputData.n_josa)
    load_ws.cell(21,3,inputData.n_tail)

    load_ws.cell(26,3,inputData.MLU_morph)
    load_ws.cell(27,3,inputData.MLU_word)
    load_ws.cell(28,3,inputData.MLU_eojeol)
    load_ws.cell(29,3,inputData.MLU_jeol)

    #simple analy
    w_row = 6
    max_index = inputData.morphDic.get_sen_morphs_max_index()
    for i in range(max_index+1):
        for morph in inputData.morphDic.get_sen_morphs(i):
            load_ws.cell(w_row, 9, w_row-5)
            draw_style(load_ws, w_row, 9)
            t, m = inputData.morphDic.get_morph(morph)
            load_ws.cell(w_row, 10, t)
            draw_style(load_ws, w_row, 10)

            load_ws.cell(w_row, 11, m)
            draw_style(load_ws, w_row, 11)

            w_row += 1
    
    #eojeul analy
    w_row = 6
    for eojeol in inputData.eojeolList:
        load_ws.cell(w_row, 12, eojeol)
        draw_style(load_ws, w_row, 12)

        w_row += 1

    w_row = 6
    temp = 0
    for m in Morph.체언:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.용언:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.관형사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.부사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.감탄사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.조사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.선어말어미:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.어말어미:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.접두사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.접미사:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.어근:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.부호:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.분석불능:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    temp = 0
    for m in Morph.한글이외:
        c = inputData.morphDic.get_morph_cnt(m)
        load_ws.cell( w_row ,17 ,c)
        temp += c
        w_row += 1
    load_ws.cell( w_row-1 ,18 ,temp)

    ## 문장
    w_row = 6
    for i, s in enumerate(inputData.sentenceList):
        load_ws.cell(w_row, 20, i+1)
        draw_style(load_ws, w_row, 20)

        load_ws.cell(w_row, 21, s)
        draw_style(load_ws, w_row, 21, option=1)

        load_ws.cell(w_row, 22, len(inputData.morphDic.get_sen_morphs(i)))
        draw_style(load_ws, w_row, 22)

        load_ws.cell(w_row, 23, len(posTagger.splitEojeol(s)))
        draw_style(load_ws, w_row, 23)

        cnt = 0
        cnt += inputData.morphDic.get_morph_sen_cnt("VV", i)
        cnt += inputData.morphDic.get_morph_sen_cnt("VA", i)
        cnt += inputData.morphDic.get_morph_sen_cnt("VCP", i)
        cnt += inputData.morphDic.get_morph_sen_cnt("VCN", i)
        load_ws.cell(w_row, 24, cnt)
        draw_style(load_ws, w_row, 24)
        
        load_ws.cell(w_row, 25, inputData.morphDic.get_morph_sen_cnt("ECE", i))
        draw_style(load_ws, w_row, 25)
        load_ws.cell(w_row, 26, inputData.morphDic.get_morph_sen_cnt("ECD", i))
        draw_style(load_ws, w_row, 26)
        load_ws.cell(w_row, 27, inputData.morphDic.get_morph_sen_cnt("ECS", i))
        draw_style(load_ws, w_row, 27)
        load_ws.cell(w_row, 28, inputData.morphDic.get_morph_sen_cnt("ETN", i))
        draw_style(load_ws, w_row, 28)
        load_ws.cell(w_row, 29, inputData.morphDic.get_morph_sen_cnt("ETD", i))
        draw_style(load_ws, w_row, 29)

        w_row += 1
    

    # option
    for i, ch in enumerate(checkOptions):
        option_morph_list = []
        if ch.get():    
            option_morph_list += Morph.get_checkbox_morph_list(i)

        for  morph in option_morph_list:
            load_wb.create_sheet(morph)
            load_sub_ws = load_wb[morph]

            load_sub_ws.cell(1,1,morph)
            load_sub_ws.cell(3,2,"번호")
            draw_style(load_sub_ws, 3, 2, option=2)
            load_sub_ws.cell(3,3,"포함문장")
            draw_style(load_sub_ws, 3, 3, option=2)
            load_sub_ws.cell(3,4,"형태소")
            draw_style(load_sub_ws, 3, 4, option=2)
            load_sub_ws.cell(3,5,"태그")
            draw_style(load_sub_ws, 3, 5, option=2)
            


            temp_index = 1
            row = 4
            morph_index = inputData.morphDic.morph_typeIndex(morph)
            sen_list = inputData.morphDic.morph_sen.get(morph_index)
            if sen_list:
                for s in sen_list:
                    morphs_sen = inputData.morphDic.get_sen_morphs(s)

                    for m in morphs_sen:
                        text, mm= inputData.morphDic.get_morph(m)
                        if mm.strip() == morph.strip():
                            load_sub_ws.cell(row, 2, temp_index)
                            draw_style(load_sub_ws, row, 2)
                            load_sub_ws.cell(row, 3, inputData.sentenceList[s])
                            draw_style(load_sub_ws, row, 3, option=1)
                            load_sub_ws.cell(row, 4, text)
                            draw_style(load_sub_ws, row, 4)
                            load_sub_ws.cell(row, 5, morph)
                            draw_style(load_sub_ws, row, 5)


                            row += 1
                            temp_index +=1


    print("completed writing report")
    return load_wb

if __name__ =="__main__":

    process = MainProcess()

    input ='''나는 학생으로서 학교생활을 하며 교칙에 묶여서 생활하는 것에 대한 불만을 품었다.
또 ‘우리가 왜 이런 교칙을 지켜야하는가?’에 대한 의문이 생겼다. 실제로 우리가 다니는 학교에는 수많은 교칙들이 있다.
대표적인 예로 ‘두발에 대한 규정’, ‘교복착용 및 단정한 모습’등이 있다. 문제가 되는 교칙들을 줄여나가는 노력이 이어지고 있지만 실제로 폐지한 곳은 많지 않다.
(노컷 뉴스)실제 대전에 있는 한 고등학교에서는 학생들이 자발적으로 교칙에 대하여 폐지를 요구하였지만 받아들여지지 않았다.
나는 이렇게 학교에서 인권을 침해하고 있다고 생각한다. 또 인권을 침해하는 교칙과 제도가 폐지되어야한다고 생각한다. 첫 번째, ‘두발 규정‘과 ’교복착용 및 단정한 모습‘이라는 교칙의 인권침해와 폐지되어야 하는 이유에 대하여 소개하여 보겠다.
이 두 교칙에 대하여 인권침해의 소지가 있다는 논란은 많이 있었다. 하지만 교칙들은 현재도 남아있다. 교칙들이 인권을 침해한다는 증거로는 세계 인권 선언에 “제23조 내가 원하는 일을 자유롭게 할 수 있다.”(네이버 지식백과)라는 조항에 완전히 어긋난다.
이 조항을 보면 말 그대로 인간들은 자신이 원하는 일을 자유로이 할 수 있다. 학생들도 마찬가지이다. 하지만 교칙들은 학생들은 틀 안에 가두고 자신의 생각과 마음을 표출하지 못하도록 하고 있다고 생각한다.
또 학생들의 개성과 취향을 존중해 주지 못한다. (노컷뉴스)그리고 이러한 교칙들로 인해 한 기사에서는 ’지난해 대전청소년인권네트워크 조사에서 대전지역 고등학생의 54%가 \'인권이 침해당하고 있다\'고 응답한 바 있다.‘라고 한다.
두 번째, 내신제도로 인한 인권침해이다. 우리나라의 내신제도는 학생들을 교실 안에만 가두어 정확한 목적을 알지  못한 채 그저 공부만 하도록 강요하고 있다.
그리고 학생들은 그저 수동적으로 수업만을 들을 뿐 능동적으로 질문을 하거나 자신의 생각을 표현할 기회가 거의 없다. 선생님들 또한 자신이 진정으로 가르치고 싶은 것은 제쳐두고 국가나 시에서 원하는 교육을 시켜주어야 한다.
실제 내신제도에 대하여 비판을 하는 책에서는 선생님인 작가와 그의 동료는 수업을 할 때마다 죄책감을 느낀다고 한다. 이러한 문제점들 또한 세계인권 선언 “19조 생각하고 표현하는 것은 자유이다.”라는 것에 어긋난다.
(네이버 지식백과)그러므로 우리나라의 내신은 학생들이 경쟁에서 벗어나 자신의 생각을 자유롭게 표현하고 선생님들은 자신이 진정 가르치고 싶은 것들을 가르칠 수 있도록 폐지되어야 한다고 생각된다.
학생들의 인권을 침해하는 ‘두발규정’과 ‘교복착용과 단정한 모습’, 학생과 선생님의 인권을 침해하는 ‘내신제도’는 수정되거나 폐지되어야 마땅하다.
이러한 교칙과 제도의 폐지를 통하여 학생들은 자신의 개성과 생각을 마음껏 드러내며 자신의 목적을 찾아내고 선생님들은 자신이 원하는 수업을 학생들에게 가르치며 진정한 스승이 될 수 있을 것이다.'''

    process.analyze(input)

    write_report(process.inputData)